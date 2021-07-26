"""
================================================================================
Description:    
    Program for determination of the cumulative distribution parameters Dv(x),
    which can be used to describe the distribution. For example, if the
    Dv(90) = 844 um, this means that 90% of the sample has a size of 844 um
    or smaller. An additional parameter to show the width of the size
    distribution is the span. The span of a volume-based size distribution is
    defined as Span = (Dv(90) – Dv(10))/Dv(50) and gives an indication of how
    far the 10 percent and 90 percent points are apart, normalized with 
    the midpoint.
Author:
    Dmitry Safonov
Organizations:
    1. Lappeenranta-Lahti University of Technology LUT (Solid/Liquid Separation
    Research Group)
    2. Metso-Outotec Oy 
Version: 
    0.2 23.07.2021
================================================================================
"""

# Standard includes:
from os.path import dirname, abspath, isfile
import sys
import numpy as np
from openpyxl import load_workbook


class DvCalculator():
    """Class of the Dv calculator"""
    def __init__(self):
        """Constructor of the class"""
        self.inputFileName = None  # Name of the file with input data
        self.outputFileName = None  # Name of the file with output data
        self.xlsxFileName = None  # Name of the xlsx file name with PSD data
        self.xlsxDataStartRow = None  # Start row of the PSD data in xlsx file
        self.xlsxDataEndRow = None  # End row of the PSD data in xlsx file
        self.xDvValues = None  # List of x values for which Dv(x) should be determined
        self.DvValues = None  # List of calculated Dv(x) values (diameters)
        self.span = None  # Span of the distribution
        self.channelsNum = None  # Number of PSD channels
        self.CEDiam_chLower = None  # Lower channel numbers of CE diameter distribution
        self.CEDiam_chCentre = None  # Center channel numbers of CE diameter distribution
        self.CEDiam_chUpper = None  # Upper channel numbers of CE diameter distribution
        self.CEDiam_diff = None  # CE diameter differential distribution values
        self.CEDiam_cum = None  # CE diameter cumulative distribution values
        
        # Put parameters default values:
        self.put_default_data()
        
    def put_default_data(self):
        """Method for putting the default values and parameters"""
        self.inputFileName = "Input_data.txt"
        self.outputFileName = "Output_data.txt"
        self.xlsxDataStartRow = 20
        self.xlsxDataEndRow = 120
        self.xDvValues = []
        self.DvValues = []
        self.span = 0
        self.channelsNum = 100
        self.CEDiam_chLower = np.zeros(100)
        self.CEDiam_chCentre = np.zeros(100)
        self.CEDiam_chUpper = np.zeros(100)
        self.CEDiam_diff = np.zeros(100)
        self.CEDiam_cum = np.zeros(100)
           
    def load_distr_data(self):
        """Method for loading the xlsx file with distribution data"""
        # Read the data from input file
        print("Reading Input_data.txt file...")
        fileName = dirname(abspath(__file__)) + "\\" + self.inputFileName
        try:
            with open(fileName, "r") as f:
                # Read file name of the xlsx file with PSD data (without \n)
                self.xlsxFileName = f.readline()[:-1]
                
                # Read x values of Dv(x)
                line = f.readline()
                while line:
                    temp = float(line)
                    # x value should be in range [0.0 - 100.0]
                    if (temp < 0.0) or (temp > 100.0):
                        raise ValueError
                    else:    
                        self.xDvValues.append(temp)
                    line = f.readline()
                f.close()
        except OSError:
            self.print_error_and_exit("Error: Input_data.txt not found!")
        except ValueError:
            self.print_error_and_exit("Error: Wrong data in Input_data.txt file!")  
        print("Done!")
        
        # Read data from desired xlsx file
        print("Reading xlsx file...")
        if isfile(self.xlsxFileName):
            # Try to open the workbook
            try:
                workbook = load_workbook(filename=self.xlsxFileName)
            except:    
                self.print_error_and_exit("Error: Can't open the xlsx file!")
                
            # Try to read the data
            try:     
                worksheet = workbook["Data"]
                ptr = 0  # pointer to the array element
                for i in range(self.xlsxDataStartRow, self.xlsxDataEndRow):
                    # Read channels data for CEDiameter
                    self.CEDiam_chLower[ptr] = float(worksheet.cell(i, 3).value)
                    self.CEDiam_chCentre[ptr] = float(worksheet.cell(i, 4).value)
                    self.CEDiam_chUpper[ptr] = float(worksheet.cell(i, 5).value)
                
                    # Read diff distribution data
                    self.CEDiam_diff[ptr] = float(worksheet.cell(i, 6).value)
                    ptr += 1
            except:
                workbook.close()
                self.print_error_and_exit("Error: Wrong data in the xlsx file!")
            workbook.close()
        else:
            self.print_error_and_exit("Error: xlsx file with PSD data not found!")
        print("Done!")

    def calculate_cum_distribution(self):
        """Method for calculating the cumulative distribution"""
        print("Calculating cumulative distribution...")
        for i in range(1, self.channelsNum):
            self.CEDiam_cum[i] = self.CEDiam_cum[i - 1] + self.CEDiam_diff[i]
        print("Done!")

    def calculate_single_Dv(self, xValue):
        """Method for calculating a single Dv(x) value"""
        xHight = 0  # Hight value of the current channel x parameter (in %) 
        xLow = 0  # Low value of the current channel x parameter (in %)
        dHight = 0  # Hight value of the current channel d parameter (in um) 
        dLow = 0  # Low value of the current channel d parameter (in um)
        # Determine the channel and channel parameters in which xValue is located
        for channel in range(self.channelsNum):
            xLow = xHight
            xHight = self.CEDiam_cum[channel]
            
            # If both xLow and xHight are 0, then continue
            if (xLow <= 0E-6) and (xHight <= 0E-6):
                continue
            
            if (xValue >= xLow) and (xValue <= xHight):
                dLow = self.CEDiam_chLower[channel]
                dHight = self.CEDiam_chUpper[channel]
                break
            
        # Calculate the Dv(x) value and return
        Dv = (xValue - xLow) * (dHight - dLow) / (xHight - xLow) + dLow
        return Dv
    
    def calculate_all_Dv_values(self):
        """Method for calculating the Dv values"""
        # Procedure for determination of all Dv(x) values
        print("Calculating Dv(x) values...")
        for xValue in self.xDvValues:
            Dv = self.calculate_single_Dv(xValue)
            self.DvValues.append(Dv)
        print("Done!")
       
    def calculate_span(self):
        """Method for calculating the span of the distribution"""
        Dv10 = self.calculate_single_Dv(10.0)
        Dv50 = self.calculate_single_Dv(50.0)
        Dv90 = self.calculate_single_Dv(90.0)
        self.span = (Dv90 - Dv10) / Dv50
    
    def write_calculated_data(self):
        """Method for writing the calculated data to output file"""
        print("Writing output data to Output_data.txt file...")
        fileName = dirname(abspath(__file__)) + "\\" + self.outputFileName
        with open(fileName, "w") as f:
            f.write("Calculation results:\n")
            ptr = 0  # Index
            for xValue in self.xDvValues:
                s = "Dv({0:.3f}) = {1:.3f} um\n".format(xValue, self.DvValues[ptr])
                f.write(s)
                ptr += 1
            f.write("Span = {0:.3f}\n".format(self.span))
            f.close()
        print("Done! Calculations have finished")
        
    def print_error_and_exit(self, text):
        """Method for printing an error and exiting the program"""
        print(text)
        sys.exit()
        
        
if __name__ == '__main__':
    # Main routine
    def main():
        calc = DvCalculator()
        calc.load_distr_data()
        calc.calculate_cum_distribution()
        calc.calculate_all_Dv_values()
        calc.calculate_span()
        calc.write_calculated_data()
    main()
